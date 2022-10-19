from builtins import print
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


class Gramatica:
    def __init__(self, Vt, Vn, S, P):
        # Terminales
        self._Vt = Vt
        # No terminales
        self._Vn = Vn
        # Simbolo inicial
        self._S = S
        # Producciones
        self._P = P
        self._primeros = {}
        self._siguientes = {}
        self._isLL1 = False
        self._conjuntoP = []
        self._tabla = []

    def setVt(self, vt):
        self._Vt = vt

    def setVn(self, vn):
        self._Vn = vn

    def setS(self, s):
        self._S = s

    def setP(self, p):
        self._P = p

    def getVt(self):
        return self._Vt

    def getVn(self):
        return self._Vn

    def getS(self):
        return self._S

    def getP(self):
        return self._P

    def getTabla(self):
        return self._tabla

    def getConjuntoP(self):
        return self._conjuntoP

    def getPrimeros(self):
        return self._primeros

    def getSiguientes(self):
        return self._siguientes

    def actualizarProduccion(self, llave, valor):
        self._P[llave] = valor

    def actualizarProduccion_nueva(self, llave, valor):
        self._P[llave] = valor + '|λ'

    def setLL1(self, isLL1):
        self._isLL1 = isLL1

    def eliminarRecursion(self):
        # A+ BIA' | BA'|  | BA'
        # A' → QA’ | Q2A' |   | AmAla

        terminos_independientes = []
        terminos_asociativos = []
        llave_prima = ''
        producciones_nuevas = []
        termino_aparte = ''

        for llave, valor in self._P.items():
            terminos_asociativos.clear()
            terminos_independientes.clear()
            hay_recursion = False

            if valor.count("|") > 0:  # Si contiene el separador  |
                valores = valor.split("|")  # Se obtiene cada conjunto
                for val in valores:
                    if val.startswith(llave):
                        print("Se encontro recursion : ", llave, " --> ", val)
                        terminos_asociativos.append(val.removeprefix(llave) + llave + '\'')
                        hay_recursion = True
                    else:
                        terminos_independientes.append(val + llave + '\'')

                if hay_recursion:
                    if len(terminos_independientes) > 0:
                        terminos = ''
                        for termino in terminos_independientes:
                            terminos += termino + '|'
                        terminos = terminos.removesuffix('|')
                        self.actualizarProduccion(llave, terminos)
                    if len(terminos_asociativos) > 0:
                        terminos = ''
                        for termino in terminos_asociativos:
                            terminos += termino + '|'
                        terminos = terminos.removesuffix('|')
                        producciones_nuevas.append((
                            llave + '\'', terminos
                        ))

            else:
                if valor.startswith(llave):
                    print("Se encontro recursion : ", llave, " --> ", valor)
                    llave_prima = llave + '\''
                    hay_recursion = True
                    termino_aparte = valor.removeprefix(llave)
                if hay_recursion:
                    self.actualizarProduccion(llave, llave_prima)
                    producciones_nuevas.append((
                        llave + '\'', termino_aparte + llave + '\''
                    ))

        if len(producciones_nuevas) > 0:
            for pn in producciones_nuevas:
                self.actualizarProduccion_nueva(pn[0], pn[1])
            print("****")
        print("--------------------------------------")

    """
    Formula general
    X→ y1,y2,y3 ..yk
    y1 → VT 6 VN
    prim(x) → si y1 es terminal, entonces agregar y1 a prim(x)
              si y1 no es terminal, entonces agregar prim(yl) a prim(x)
              si y1 es , entonces agregar prim(y2) a prim(x)
              si y1 hasta yk tiene 1, entonces agregar a prim(x)
    """

    def generarPrimeros(self):
        print("Generando primeros ")
        primerosTotales = {}
        # Recorremos cada produccion
        print(self._P.items())
        for llave, valor in self._P.items():
            primeros = []
            # Recorremos cada termino
            if valor.count("|") > 0:
                valores = valor.split("|")

                # Se obtiene cada conjunto
                print("Valores : ", valores)
                print("No terminales : ", self._Vt)
                for termino in valores:
                    print("Termino que se esta evaluando : ", termino[0])

                    if termino in self._Vt:
                        print(termino, "Es terminal !")
                        primeros.append(termino)

                    elif termino in self._Vn:
                        print(termino, "Es no terminal !")
                        primeros.append(self.primeroXTermino(termino, primeros))

                    if termino[0] in self._Vn:
                        print(termino[0], "Es no terminal !")
                        primeros = self.primeroXTermino(termino[0])
                        # si y1 no es terminal, entonces agregar prim(yl) a prim(x)
                    elif termino[0] in self._Vt:
                        print(termino[0], "Es terminal !")
                        if termino[0] not in primeros:
                            primeros.append(termino[0])
                    elif termino[0] == 'λ':
                        print(termino[0], 'Encontramos un lambda')
                        if len(termino) == 1:
                            primeros.append(termino[0])
                        else:
                            primeros.append(self.primeroXTermino(termino[1]))
            else:
                print("Termino que se esta evaluando : ", valor[0])

                if valor in self._Vt:
                    print(valor, "Es terminal !")
                    primeros.append(valor)

                elif valor in self._Vn:
                    print(valor, "Es no terminal !")
                    primeros.append(self.primeroXTermino(valor, primeros))

                if valor[0] in self._Vn:
                    print(valor[0], "Es no terminal !")
                    primeros = self.primeroXTermino(valor[0], [], {})
                    # si y1 no es terminal, entonces agregar prim(yl) a prim(x)
                elif valor[0] in self._Vt:
                    print(valor[0], "Es terminal !")
                    if valor not in primeros:
                        primeros.append(valor[0])
                elif valor[0] == 'λ':
                    print(valor[0], 'Encontramos un lambda')
                    if len(valor) == 1:
                        primeros.append(valor[0])
                    else:
                        primeros.append(self.primeroXTermino(valor[1]))
            print('Primeros de ', llave, '--> prim(', llave, ') = ', primeros)
            print("-----------------------------")
            primerosTotales[llave] = primeros
        self._primeros = primerosTotales

    def primeroXTermino(self, termino, primeros, total_siguientes):
        # Ingrese un termino no terminal hasta que me devuelva una lista con sus terminales
        # Recorremos las producciones para  buscar la que empieza con el termino, la coincidencia
        # {'S': 'aB|bA', 'A': 'B|aS|bAA', 'B': 'b|bS|aBB'}
        # dict_items([('E', "TE'"), ("E'", "+TE'|λ"), ('T', "FT'"), ("T'", "*FT'|λ"), ('F', 'id|(E)')])
        for no_terminal, terminos in self._P.items():
            if no_terminal == termino:
                # Recorremos sus terminos
                if terminos.count("|") > 0:
                    ter = terminos.split("|")
                    for t in ter:
                        print("Termino que se esta evaluando : ", t)
                        print("Termino que se esta evaluando : ", t[0])

                        if t in self._Vt:
                            print(t, "Es terminal !")
                            primeros.append(t)

                        elif t in self._Vn:
                            print(t, "Es no terminal !")
                            primeros.append(self.primeroXTermino(t, primeros, total_siguientes))


                        elif t == 'λ' or t[0] == 'λ':
                            print('Encontramos un lambda')
                            simbolos_a_agregar = self.siguienteXTermino(no_terminal, total_siguientes, [])
                            for s in simbolos_a_agregar:
                                if s not in primeros:
                                    primeros.append(s)
                        if t[0] in self._Vt:
                            print(t[0], "Es terminal !")
                            primeros.append(t[0])

                        elif t[0] in self._Vn:
                            print(t[0], "Es no terminal !")
                            primeros.append(self.primeroXTermino(t[0], primeros, total_siguientes))

                else:
                    print("Termino que se esta evaluando : ", terminos[0])

                    if terminos[0] in self._Vt:
                        print(terminos[0], "Es terminal !")
                        primeros.append(terminos[0])
                    elif terminos[0] in self._Vn:
                        print(terminos[0], "Es no terminal !")
                        self.primeroXTermino(terminos[0], primeros, total_siguientes)
                    elif terminos[0] == 'λ':
                        print('Encontramos un lambda')
                        simbolos_a_agregar = self.siguienteXTermino(no_terminal, total_siguientes, [])
                        for s in simbolos_a_agregar:
                            if s not in primeros:
                                primeros.append(s)
            elif termino in self.getVt():
                return termino
        return primeros

    """
    A --> αχβ
    Sig(x) -->    
    * si x es la producion inicial agregar $ a los Sig(x)
    * si ß es terminal, entonces agregar B a Sig(x)
    * si ß no es terminal entonces agregar a Sig(x) los prim(ß)
    * si ß es λ, entonces agregar a Sig(x) los sig(A)
    """

    def generarSiguientes(self):
        print("Generando siguientes ")
        print(self._P.items())
        llaves = list(self._P.keys())
        valores = list(self._P.values())

        total_siguientes = {}
        # Recorremos todas las producciones
        for llave_x, valor_x in self._P.items():
            siguientes = []
            # Recorremos todos los valores de las producciones para buscar a llave
            for llave_y, valor_y in self._P.items():
                if llave_x in valor_y:
                    division = valor_y.split(llave_x)
                    caracter_posterior = division[len(division) - 1]

                    # switch del caracter que encontro
                    if caracter_posterior == " " or caracter_posterior == "":
                        print("No hay nada despues de ", llave_x, " en ", valor_y)
                        if llave_y in total_siguientes.keys():
                            for i in total_siguientes[llave_y]:
                                if i not in siguientes:
                                    siguientes.append(i)
                        else:
                            siguientes.append(self.siguienteXTermino(caracter_posterior, total_siguientes, []))
                    elif caracter_posterior in self._Vt:
                        if caracter_posterior not in siguientes:
                            siguientes.append(caracter_posterior)
                    # Si es no terminal
                    elif caracter_posterior in self._Vn:
                        simbolos_a_agregar = self.primeroXTermino(caracter_posterior, [], total_siguientes)
                        for s in simbolos_a_agregar:
                            if s not in siguientes:
                                siguientes.append(s)
                    elif '\'' in division:
                        continue
                # Valida si la llave, valor es de la primer posicion
            # if self._P.items().index((llave_x, valor_x)) == 0:

            print("llaves :", llaves)
            print("valores :", valores)
            print(llave_x)
            print(valor_x)
            if llave_x == llaves[0] and valor_x == valores[0]:
                siguientes.append('$')

            print('Siguientes de ', llave_x, '--> sig(', llave_x, ') = ', siguientes)
            total_siguientes[llave_x] = siguientes
            print('------------------------------------------------------------------')
        self._siguientes = total_siguientes

    def siguienteXTermino(self, termino, total_siguientes, siguientes):
        for llave_y, valor_y in self._P.items():
            if termino in valor_y:
                division = valor_y.split(termino)
                caracter_posterior = division[len(division) - 1]
                if caracter_posterior == " " or caracter_posterior == "":
                    print("No hay nada despues de ", termino, " en ", valor_y)
                    ######################################################implementar
                    if llave_y in total_siguientes.keys():
                        for i in total_siguientes[llave_y]:
                            if i not in siguientes:
                                siguientes.append(i)
                    else:
                        siguientes.append(self.siguienteXTermino(caracter_posterior, total_siguientes, []))
                elif caracter_posterior in self._Vt:
                    if caracter_posterior not in siguientes:
                        siguientes.append(caracter_posterior)
                elif caracter_posterior in self._Vn:
                    simbolos_a_agregar = self.primeroXTermino(caracter_posterior, [])
                    for s in simbolos_a_agregar:
                        if s not in siguientes:
                            siguientes.append(s)
                elif '\'' in division:
                    continue
        return siguientes

        """
    
        if len(regla) != 0:
            if regla[0] in list(gramatica.keys()):
                lista = []
                der_reglas = gramatica[regla[0]]
                for itr in der_reglas:
                    indivRes = primero(itr)
                    if type(indivRes) is list:
                        for i in indivRes:
                            lista.append(i)
                    else:
                        lista.append(indivRes)
    
                if '#' not in lista:
                    return lista
                else:
                    nuevaLista = []
                    lista.remove('#')
                    if len(regla) > 1:
                        nuevaRes = primero(regla[1:])
                        if nuevaRes != None:
                            if type(nuevaRes) is list:
                                nuevaLista = lista + nuevaRes
                            else:
                                nuevaLista = lista + [nuevaRes]
                        else:
                            nuevaLista = lista
                        return nuevaLista
                    lista.append('#')
                    return lista
        """

    def conjuntoPrediccion(self):
        print("Generando conjunto prediccion ")
        """Si todos dan 0 nos encontramos ante un conjunto de producciones que pertenecen a la gramatica LL1"""
        # Primero debemos tener los conjuntos de primeros y siguientes para aplicar la formular
        producciones = list(self.getP().keys())
        conjunto_prediccion = []

        for produccion, terminos in self.getP().items():
            valores = terminos.split("|")

            repetidos = 0
            # print(valores)
            for v in valores:
                primeros = []
                if v != "":
                    primero = self.primeroXTermino(v[0], [], self.getSiguientes())
                    if not primero:
                        primero = self.primeroXTermino(v, [], self.getSiguientes())
                    if primero:
                        if type(primero) == list:
                            for p in primero:
                                if p in primeros:
                                    repetidos = 1
                                primeros.append(p)
                        else:
                            if primero in primeros:
                                repetidos = 1
                            primeros.append(primero)
                if v == 'λ':
                    simbolos_a_agregar = self.siguienteXTermino(produccion, self.getSiguientes(), [])
                    for s in simbolos_a_agregar:
                        if s not in primeros:
                            primeros.append(s)
                print("conjunto prediccion de ", produccion, " --> ", v, primeros)
                conjunto_prediccion.append((produccion, primeros))

        for conjunto in conjunto_prediccion:
            print(conjunto)
        print('--------------------------------------------------------------')
        self._conjuntoP = conjunto_prediccion
        if repetidos:
            print("NO nos encontramos ante un conjunto de producciones que pertenecen a la gramatica LL1")
        else:
            print("Nos encontramos ante un conjunto de producciones que pertenecen a la gramatica LL1")
            self.setLL1(True)
        print('--------------------------------------------------------------')

    def guardarTabla(self):
        print("Guardando tabla ...")
        wb = Workbook()
        ws = wb.active
        ws.title = "TABLA ANALISIS SINTACTICO"
        ws.append(['TABLA ANALISIS SINTACTICO'])

        fuente = Font(name='Calibri',
                      size=16,
                      bold=False,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000')
        alignment = Alignment(horizontal='center',

                              vertical='center',

                              text_rotation=0,

                              wrap_text=False,

                              shrink_to_fit=False,

                              indent=0)
        for row in self.getTabla():
            ws.append(row)

        c = ws['A1:A8']
        for a in c:
            a[0].font = fuente
            a[0].alignment = alignment
        k = ws['B3:B8']
        k = ws['B3:B8']
        k = ws['B3:B8']
        for a in k:
            a[0].font = fuente
            a[0].alignment = alignment
        ws.rows
        wb.save('src/TABLA ANALISIS SINTACTICO.xlsx')

    """
    Se llena la tabla mirando el conjunto prediccion 
    se busca la fila donde este el que produce y en cada union
    con cada columna donde este cada elemento de este conjunto se coloca
    la produccion total
    ej. E --> TE', tiene conjunto prediccion {id, (}
    Entonces en la fila [E, id] Se coloca  E --> TE' y en 
    la fila [E, (] tambien Se coloca  E --> TE'        
    """

    def generarTabla(self):
        print("Generando tabla ")
        columnas = []  # terminales
        filas = []  # No terminales

        for no_terminal, produccion in self.getConjuntoP():
            for termino in produccion:
                if termino not in columnas:
                    columnas.append(termino)
            if no_terminal not in filas:
                filas.append(no_terminal)

        tabla = [[0 for i in range(len(filas) + 2)] for j in range(len(columnas) + 1)]
        tabla[0][0] = "VT / VN"

        for fila in tabla:
            print(fila)

        for columna in range(1, len(tabla[0]), 1):
            tabla[0][columna] = columnas[columna - 1]

        for columna in range(1, len(tabla), 1):
            tabla[columna][0] = filas[columna - 2]

        print('--------------------------------------------------------------')

        for produccion, terminos in self.getConjuntoP():
            for t in terminos:
                if t != "":
                    for x in range(1, len(tabla)):
                        if tabla[x][0] == produccion:
                            if t in tabla[0]:
                                tabla[x][tabla[0].index(t)] = produccion + ' --> ' + self.getTerminos(produccion)

        for fila in tabla:
            print(fila)

        self._tabla = tabla

    def getTerminos(self, no_terminal):
        for nt, produccion in self.getP().items():
            if no_terminal == nt:
                return str(produccion)
